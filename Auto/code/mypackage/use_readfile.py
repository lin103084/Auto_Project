# 讀取Excel 的程式碼
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 讀取CSV 的程式碼
import csv

#使用GUI介面讀取檔案
import tkinter as tk
from tkinter import filedialog

def excel_file():
    root = tk.Tk()
    root.withdraw()
    print('選取Excel域名檔案!')
    file_path = filedialog.askopenfilename(parent=root, title='選取Excel，域名檔案!',
                                            filetypes=[("Excel files", ".xlsx .xls")])
    print(f'選取路徑 :{file_path}\n')

    wb = load_workbook(file_path)
    ws = wb.active
    ws_list = []
    for row in range(1,999):
        for col in range(1,999):
            char = get_column_letter(col)
            if ws[char + str(row)].value == None:
                continue
            else:
                ws_list.append(ws[char + str(row)].value)
    return ws_list


def csv_file():
    root = tk.Tk()
    root.withdraw()
    print('選取csv域名檔案!')
    file_path = filedialog.askopenfilename(parent=root, title='選取csv，域名檔案!',
                                            filetypes=[("CSV files", ".csv")])
    print(f'選取路徑 :{file_path}\n')

    csv_list = []
    with open(file_path, newline='', encoding= 'utf-8-sig') as csvfile:
        # 讀取 CSV 檔案內容
        rows = csv.reader(csvfile)
        # 以迴圈輸出每一列
        for row in rows:
            csv_list.append(row)

    return csv_list



if __name__ == '__main__':
    excel_file()
    csv_file()